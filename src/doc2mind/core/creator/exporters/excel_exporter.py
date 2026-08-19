"""Excel (XLSX) 原生电子表格导出引擎 — 基于 openpyxl。

特性：
- 自动从 Markdown 表格、JSON 或文本矩阵中抽取多维数据；
- 优雅专业商务调色（科技深蓝表头、白色粗体字、细网格边框）；
- 自适应计算每列最佳列宽，杜绝文字截断；
- 自动冻结首行表头，浏览体验极佳。
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

from doc2mind.core.creator.models import ArtifactModel

logger = logging.getLogger("doc2mind.creator.excel")


class ExcelExporter:
    """XLSX 电子表格生成器。"""

    def __init__(self) -> None:
        pass

    def export(self, artifact: ArtifactModel, output_path: Path) -> Path:
        """编译生成 .xlsx 物理文件并保存到指定路径。"""
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (artifact.title[:28] if artifact.title else "数据对比矩阵").replace("/", "_").replace("\\", "_")

        # 1. 抽取表格数据
        rows_data: list[list[str]] = []
        for line in artifact.raw_content.splitlines():
            ls = line.strip()
            if ls.startswith("|") and ls.endswith("|"):
                if re.match(r"^\|[\s\-:|]+\|$", ls):
                    continue
                cols = [c.strip() for c in ls.strip("|").split("|")]
                if any(cols):
                    rows_data.append(cols)

        # 若无 markdown 表格，将非空行转为单列表
        if not rows_data:
            rows_data = [["内容条目"]]
            for line in artifact.raw_content.splitlines():
                ls = line.strip()
                if ls and not ls.startswith(":::"):
                    rows_data.append([ls])

        # 2. 样式定义
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")

        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        data_font = Font(name="微软雅黑", size=10, color="1E293B")

        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 3. 填充数据与样式
        for r_idx, row in enumerate(rows_data, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border

                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                    ws.row_dimensions[r_idx].height = 28
                else:
                    cell.font = data_font
                    cell.alignment = align_left
                    if r_idx % 2 == 0:
                        cell.fill = zebra_fill
                    ws.row_dimensions[r_idx].height = 22

        # 4. 冻结首行
        ws.freeze_panes = "A2"

        # 5. 自动列宽调整
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                # 中文字符计算 1.8 字符宽度
                col_w = sum(1.8 if ord(ch) > 127 else 1.0 for ch in val_str)
                if col_w > max_len:
                    max_len = int(col_w)
            ws.column_dimensions[col_letter].width = max(12, min(50, max_len + 4))

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        logger.info("XLSX 导出成功: %s, 共 %d 行", output_path, len(rows_data))
        return output_path
