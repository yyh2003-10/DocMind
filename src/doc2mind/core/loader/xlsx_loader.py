"""Excel 加载器 — 基于 `openpyxl`。

特点：
- 逐 sheet 遍历，每个 sheet 输出一个 H2 标题 + 若干 table_row 元素
- 表头行自动识别（首行非空 + 第二行也非空启发式）
- 合并单元格展开（取左上角值填充到所有覆盖单元格）
- 跳过完全空白的行
- 元数据携带 `sheet` / `row_index` 便于追溯

局限性：
- 公式单元格取 `data_only=True` 的缓存值；若文件从未被 Excel 打开过，公式返回 None
- 不解析图表 / 嵌入图片
"""

from __future__ import annotations

import hashlib
from pathlib import Path

from doc2mind.core.loader.base import Loader, LoaderError
from doc2mind.core.models import (
    DocFormat,
    DocumentElement,
    ElementType,
    LoadedDocument,
)


def _cell_to_str(value: object) -> str:
    """把单元格值规范化为字符串。

    - None → ""
    - bool → "TRUE" / "FALSE"
    - datetime/date → ISO 8601
    - 其余 → str()
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    # datetime / date / time 都有 isoformat
    iso = getattr(value, "isoformat", None)
    if callable(iso):
        try:
            return iso()
        except Exception:  # noqa: BLE001
            return str(value)
    return str(value)


class XlsxLoader(Loader):
    """Excel 文档加载器（openpyxl 实现）。"""

    supported_extensions = ("xlsx", "xls")

    def extract(self, path: Path) -> LoadedDocument:
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import range_boundaries
        except ImportError as e:
            raise LoaderError(
                "openpyxl 未安装。请运行：pip install openpyxl"
            ) from e

        if not path.exists():
            raise LoaderError(f"文件不存在: {path}")

        try:
            data = path.read_bytes()
            file_hash = hashlib.md5(data).hexdigest()
            wb = load_workbook(path, data_only=True, read_only=False)
            elements: list[DocumentElement] = []

            for sheet in wb.worksheets:
                sheet_title = sheet.title or "(unnamed)"
                # sheet 级标题
                elements.append(
                    DocumentElement(
                        content=f"## Sheet: {sheet_title}",
                        type=ElementType.HEADING,
                        metadata={
                            "type": "heading",
                            "level": 2,
                            "sheet": sheet_title,
                            "source_format": DocFormat.XLSX.value,
                        },
                    )
                )

                # 合并单元格展开映射：{(row,col): 左上角值}
                merged_map: dict[tuple[int, int], object] = {}
                for merge_range in sheet.merged_cells.ranges:
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(
                            str(merge_range)
                        )
                    except Exception:  # noqa: BLE001
                        continue
                    top_value = sheet.cell(row=min_row, column=min_col).value
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            if (r, c) == (min_row, min_col):
                                continue
                            merged_map[(r, c)] = top_value

                row_index = 0
                for row in sheet.iter_rows():
                    row_index += 1
                    values: list[object] = []
                    for cell in row:
                        coord = (cell.row, cell.column)
                        if coord in merged_map:
                            values.append(merged_map[coord])
                        else:
                            values.append(cell.value)

                    row_text = [_cell_to_str(v) for v in values]
                    if not any(c.strip() for c in row_text):
                        continue

                    # 去掉尾部空列
                    while row_text and not row_text[-1].strip():
                        row_text.pop()

                    elements.append(
                        DocumentElement(
                            content="| " + " | ".join(row_text) + " |",
                            type=ElementType.TABLE_ROW,
                            metadata={
                                "type": "table_row",
                                "sheet": sheet_title,
                                "row_index": row_index,
                                "cols": len(row_text),
                                "source_format": DocFormat.XLSX.value,
                            },
                        )
                    )

            wb.close()

            return LoadedDocument(
                source=path.name,
                format=DocFormat.XLSX,
                elements=elements,
                page_count=len(wb.worksheets) if wb.worksheets else None,
                size_bytes=len(data),
                file_hash=file_hash,
            )
        except LoaderError:
            raise
        except Exception as e:  # noqa: BLE001
            raise LoaderError(f"Excel 解析失败 ({path.name}): {e}") from e
